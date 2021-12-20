from math import ceil
import os
from datetime import datetime, timedelta
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
import yaml
import download


def get_class(note: str):
    if note == 'Utilizzo Generico':
        return ''

    note = note.replace('VBPS030009 - ', '')
    cls = note.split()[0]
    return cls.strip()


def get_surname(complete_name: str):
    words = map(lambda x: x.lower(), complete_name.strip().split())
    word: str
    surname_parts = [word.capitalize() for word in words if word in surnames]
    return ' '.join(surname_parts)


with open('/config/config.yaml') as file:
    config = yaml.safe_load(file)

with open(config['surnames_path']) as file:
    surnames = [line.strip() for line in file.readlines()]

'''
user_home = os.path.expanduser('~')

download_folder = os.path.join(user_home, 'Downloads')
'''
download_folder = config['download_path']

filename_prefix = datetime.now().strftime('%d%m%Y')

files = os.listdir(download_folder)

filename = [f for f in files if re.match(f'export_{filename_prefix}.', f)][0]

print(f'Loading {filename}')
filepath = os.path.join(download_folder, filename)

with open(filepath, 'r') as file:
    lines = file.readlines()

lines = [line.strip() for line in lines]

lines = lines[lines.index('<Row>') + 1:]

parsed_lines = []

row = 0
while row < len(lines):
    new_line = []
    while '<Cell>' in lines[row]:
        line = lines[row]

        line = line.replace('<Cell><Data ss:Type="String">', '')
        line = line.replace('</Data></Cell>', '')
        new_line.append(line)
        row += 1
    if new_line:
        parsed_lines.append(new_line)
    row += 1

df = pd.DataFrame(data=parsed_lines[1:], columns=parsed_lines[0])

output_name = os.path.join(config['output_path'], 'reservations.xlsx')

workbook = Workbook()
sheet = workbook.active

sheet.merge_cells('A1:F1')
sheet.row_dimensions[1].height = 30

title_text = 'Laboratorio Informatica'
title_size = 25
title_width = len(title_text)

title = sheet.cell(row=1, column=1)
title.value = title_text
title.alignment = Alignment(horizontal='center', vertical='center')

side = Side(border_style=None)
no_border = Border(left=side, right=side, top=side, bottom=side)

side = Side(border_style='thin', color='000000')
boxed = Border(left=side, right=side, top=side, bottom=side)
title.border = no_border
title.font = Font(name='Arial', sz=25)

now = datetime.now()
weekday = now.isoweekday()
monday = now - timedelta(days=weekday - 1)
friday = monday + timedelta(days=4)

sheet.merge_cells('A2:F2')
sheet.row_dimensions[2].height = 30

week_text = f'{monday.strftime("%d/%m/%Y")} - {friday.strftime("%d/%m/%Y")}'
week_text_size = title_size
week_text_width = len(week_text)

week_row = sheet.cell(row=2, column=1)
week_row.value = week_text
week_row.alignment = Alignment(horizontal='center', vertical='center')
week_row.border = no_border
week_row.font = Font(name='Arial', sz=25)

centered = Alignment(horizontal='center', vertical='center')

weekdays = ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì']

for col, weekday in enumerate(weekdays):
    cell = sheet.cell(row=3, column=col + 2)
    cell.value = weekday
    cell.alignment = centered

bounds = [(timedelta(hours=8, minutes=0), timedelta(hours=8, minutes=50)),
          (timedelta(hours=8, minutes=50), timedelta(hours=9, minutes=40)),
          (timedelta(hours=9, minutes=40), timedelta(hours=10, minutes=40)),
          (timedelta(hours=10, minutes=40), timedelta(hours=11, minutes=30)),
          (timedelta(hours=11, minutes=30), timedelta(hours=12, minutes=30)),
          (timedelta(hours=12, minutes=30), timedelta(hours=13, minutes=40)),
          (timedelta(hours=14, minutes=0), timedelta(hours=15, minutes=0)),
          (timedelta(hours=15, minutes=0), timedelta(hours=16, minutes=0))]

lessons = ['1 8:00-8:50', '2 8:50-9:40', '3 9:50-10:40', '4 10:40-11:30', '5 11:40-12:30', '6 12:30-13:20',
           '7 14:10-15:00', '8 15:00-16:00']
sheet.column_dimensions['A'].width = max([len(lesson) for lesson in lessons])

for row, lesson in enumerate(lessons):
    cell = sheet.cell(row=row + 4, column=1)
    cell.value = lesson
    cell.alignment = centered

for idx, reservation in df.iterrows():
    weekday = datetime.strptime(reservation['data_inizio'], '%Y-%m-%d').isoweekday()
    column = weekday + 1

    h, m, _ = map(lambda x: int(x), reservation['ora_inizio'].split(':'))

    start = timedelta(hours=h, minutes=m)

    h, m, _ = map(lambda x: int(x), reservation['ora_fine'].split(':'))
    end = timedelta(hours=h, minutes=m)
    length = ceil(((end - start).total_seconds()) / 3600)

    row = 4
    while not (bounds[row - 4][0] <= start < bounds[row - 4][1]):
        row += 1

    for i in range(row, row + length):
        cls = get_class(reservation['utilizzatore'])
        if reservation['a_nome_di'] == 'DE NICOLA MICAELA MARIA':
            prof = 'De Nicola'
        else:
            prof = get_surname(reservation['a_nome_di'])
        cell = sheet.cell(i, column)
        cell.value = f'{cls} ({prof})'.strip()
        cell.alignment = centered

for col in range(6):
    max_length = max([len(sheet.cell(row, 6).value) for row in range(3, 9, 1) if sheet.cell(row, 6).value is not None])
    sheet.column_dimensions[chr(ord('A') + col)].width = max_length

for row in range(3, 12, 1):
    for col in range(1, 7):
        cell = sheet.cell(row, col)
        cell.border = boxed

Worksheet.set_printer_settings(sheet, orientation='landscape', paper_size='9')
sheet.print_options.horizontalCentered = True
sheet.print_options.verticalCentered = True

workbook.save(output_name)
